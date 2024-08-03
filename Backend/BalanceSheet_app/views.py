from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.db.models import Sum
from Expenses_app.models import *
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from io import BytesIO
from django.core.mail import EmailMessage
from daily_expense_system.settings import EMAIL_HOST_USER

class BalanceSheetEmailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # Creating a new excel sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Adding headers
        headers = ['Date', 'Name', 'Amount', 'Type', 'Note', 'Paid', 'Owed']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        personal_expenses = Expense.objects.filter(user=user, expense_type='Personal')
        group_expenses = Expense.objects.filter(user=user, expense_type='Group')

        row = 2
        total_paid = 0
        total_owed = 0

        # Add the personal expenses to the sheet
        for expense in personal_expenses:
            ws.append([
                expense.date.strftime('%Y-%m-%d'),
                expense.name,
                float(expense.amount),
                'Personal',
                expense.note,
                float(expense.amount),
                0
            ])
            total_paid += float(expense.amount)
            row += 1

        # Add the group expenses to the sheet
        for expense in group_expenses:
            group_details = expense.group_details.all()
            
            for detail in group_details:
                if detail.username == user.username:
                    ws.append([
                        expense.date.strftime('%Y-%m-%d'),
                        expense.name,
                        float(detail.amount),
                        'Group',
                        expense.note,
                        float(detail.amount) if detail.is_paid else 0,
                        0
                    ])
                    total_paid += float(detail.amount) if detail.is_paid else 0
                else:
                    ws.append([
                        expense.date.strftime('%Y-%m-%d'),
                        expense.name,
                        float(detail.amount),
                        'Group',
                        expense.note,
                        0,
                        float(detail.amount) if not detail.is_paid else 0
                    ])
                    total_owed += float(detail.amount) if not detail.is_paid else 0
                row += 1

        # Adding totals
        ws.cell(row=row, column=1, value="Total")
        ws.cell(row=row, column=6, value=total_paid)
        ws.cell(row=row, column=7, value=total_owed)

        # Save the excel sheet to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create the email
        subject = 'Your Balance Sheet'
        message = 'Please find attached your balance sheet.'
        email = EmailMessage(
            subject,
            message,
            EMAIL_HOST_USER,
            [user.email],
        )
        email.attach('balance_sheet.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Send the email
        try:
            email.send()
            return Response({"message": "Balance sheet has been sent to your email."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
class BalanceSheetView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        balance_sheet = {
            'personal_expenses': [],
            'group_expenses': [],
            'total_paid': 0,
            'total_owed': 0
        }

        personal_expenses = Expense.objects.filter(user=user, expense_type='Personal')
        
        for expense in personal_expenses:
            balance_sheet['personal_expenses'].append({
                'date': expense.date,
                'name': expense.name,
                'amount': float(expense.amount),
                'note': expense.note,
                'paid': float(expense.amount),
                'owed': 0
            })
            balance_sheet['total_paid'] += float(expense.amount)

        group_expenses = Expense.objects.filter(user=user, expense_type='Group')

        for expense in group_expenses:
            group_details = expense.group_details.all()
            
            for detail in group_details:
                if detail.username == user.username:
                    balance_sheet['group_expenses'].append({
                        'date': expense.date,
                        'name': expense.name,
                        'amount': float(detail.amount),
                        'note': expense.note,
                        'paid': float(detail.amount) if detail.is_paid else 0,
                        'owed': 0
                    })
                    balance_sheet['total_paid'] += float(detail.amount) if detail.is_paid else 0
                else:
                    balance_sheet['group_expenses'].append({
                        'date': expense.date,
                        'name': expense.name,
                        'amount': float(detail.amount),
                        'note': expense.note,
                        'paid': 0,
                        'owed': float(detail.amount) if not detail.is_paid else 0
                    })
                    balance_sheet['total_owed'] += float(detail.amount) if not detail.is_paid else 0

        return Response(balance_sheet, status=status.HTTP_200_OK)